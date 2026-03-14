from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
import os

app = Flask(__name__)
CORS(app)

FILE = "users.xlsx"

if not os.path.exists(FILE):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Username","Password"])
    wb.save(FILE)

@app.route("/api/register", methods=["POST"])
def register():

    data = request.json
    username = data["username"]
    password = data["password"]

    wb = openpyxl.load_workbook(FILE)
    sheet = wb.active
    sheet.append([username,password])
    wb.save(FILE)

    return jsonify({"message":"saved"})


@app.route("/api/login", methods=["POST"])
def login():

    data = request.json
    username = data["username"]
    password = data["password"]
print(username,password)
    wb = openpyxl.load_workbook(FILE)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if row[0] == username and row[1] == password:
            return jsonify({"token":"ok"})

    return jsonify({"message":"invalid"})


if __name__ == "__main__":
    app.run(port=5000)